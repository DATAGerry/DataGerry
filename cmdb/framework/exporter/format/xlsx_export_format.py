# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C)  becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""
Implementation of XlsxExportFormat
"""
import logging
import json
import re
import tempfile
from openpyxl import Workbook

from cmdb.framework.exporter.format.base_exporter_format import BaseExporterFormat
from cmdb.framework.exporter.config.exporter_config_type_enum import ExporterConfigType
from cmdb.framework.rendering.render_result import RenderResult
# -------------------------------------------------------------------------------------------------------------------- #

LOGGER = logging.getLogger(__name__)

# -------------------------------------------------------------------------------------------------------------------- #
#                                               XlsxExportFormat - CLASS                                               #
# -------------------------------------------------------------------------------------------------------------------- #
class XlsxExportFormat(BaseExporterFormat):
    """
    The XLSX export format class for exporting data to Excel (.xlsx) files

    Extends: BaseExporterFormat
    """
    FILE_EXTENSION = "xlsx"
    LABEL = "XLSX"
    MULTITYPE_SUPPORT = True
    ICON = "file-excel"
    DESCRIPTION = "Export as XLS"
    ACTIVE = True


    def export(self, data: list[RenderResult], *args) -> bytes:
        """
        Exports a list of RenderResult objects as an XLSX file

        Args:
            data (list[RenderResult]): A list of `RenderResult` objects to be exported
            *args: Optional arguments, including 'metadata' and 'view', that can customize the export

        Returns:
            bytes: The content of the XLSX file as a byte string.
        """
        workbook = self.create_xls_object(data, args)

        # Save the workbook to a temporary file and return its content as bytes
        with tempfile.NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            return tmp.read()


    def create_xls_object(self, data: list[RenderResult], args) -> Workbook:
        """
        Creates an XLSX workbook with the provided data

        Args:
            data (list[RenderResult]): List of RenderResult objects
            args (tuple): Arguments containing metadata and view settings
                The first argument should be a dictionary with optional keys:
                - "metadata" (str or None): Metadata in JSON format
                - "view" (str): The view type (e.g., 'native')

        Returns:
            Workbook: The created XLSX workbook
        """
        # Create workbook
        workbook = Workbook()

        # Remove the default sheet created by openpyxl
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        # Sort the data by type_id
        decorated = [(dict_.type_information['type_id'], dict_) for dict_ in data]
        decorated = sorted(decorated, key=lambda x: x[0], reverse=False)
        sorted_list = [dict_ for (_, dict_) in decorated]

        # Initialize values
        header = ['public_id', 'active']
        columns = [] if not data else [x['name'] for x in data[0].fields]
        view = 'native'

        # Export only the shown fields chosen by the user
        if args and args[0].get("metadata", False) and\
           args[0].get('view', 'native').upper() == ExporterConfigType.RENDER.name:
            _meta = json.loads(args[0].get("metadata", ""))
            view = args[0].get('view', 'native')
            header = _meta['header']
            columns = _meta['columns']

        # Initialize current_type_id to track sheet changes
        current_type_id = None
        row_index = 1  # Start inserting data from row 1 for the header

        # Loop through sorted data and insert rows
        for obj in sorted_list:
            # Check if we need to create a new sheet based on type_id
            if current_type_id != obj.type_information['type_id']:
                current_type_id = obj.type_information['type_id']
                sheet = workbook.create_sheet(title=self.__normalize_sheet_title(obj.type_information['type_label']))

                # Insert headers in the new sheet
                for col_index, header_item in enumerate(header, start=1):
                    sheet.cell(row=row_index, column=col_index).value = header_item

                # Insert column headers for fields
                for col_index, field_name in enumerate(columns, start=len(header) + 1):
                    sheet.cell(row=row_index, column=col_index).value = field_name

                row_index += 1  # Move to the next row after headers

            # Insert data for each object
            for col_index, header_item in enumerate(header, start=1):
                header_item = 'object_id' if header_item == 'public_id' else header_item
                sheet.cell(row=row_index, column=col_index).value = str(obj.object_information.get(header_item, ""))

            # Insert data for each field
            for col_index, field_name in enumerate(columns, start=len(header) + 1):
                field_value = self._get_field_value(obj, field_name, view)  # Using 'view' to render correctly
                sheet.cell(row=row_index, column=col_index).value = str(field_value)

            row_index += 1  # Move to the next row

        return workbook


    def _get_field_value(self, obj: RenderResult, field_name: str, view: str) -> str:
        """
        Retrieves the value for a given field from the object, using the 'view' to determine how to render it

        Args:
            obj (RenderResult): The object whose field value is to be retrieved
            field_name (str): The name of the field whose value is to be retrieved
            view (str): The view configuration to control the rendering of the value

        Returns:
            str: The field value
        """
        for field in obj.fields:
            if field['name'] == field_name:
                return BaseExporterFormat.summary_renderer(obj, field, view)  # Render based on 'view'

        return ""


    def __normalize_sheet_title(self, input_data: str) -> str:
        """
        Normalizes the sheet title by replacing invalid characters

        Args:
            input_data (str): The raw sheet title

        Returns:
            str: The normalized sheet title
        """
        return re.sub(r'[\\*?:/\[\]]', '_', input_data)
